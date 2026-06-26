"""
cma/excel_report.py
===================
Renders the CMA workbook from the shared section data in report_sections.py,
so the Excel and the PDF (pdf_report.py) are guaranteed to be replicas.
"""

from typing import Dict, Any
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .report_sections import build_sections

# ── Palette ───────────────────────────────────────────────────────────────────
_NAVY, _ACCENT, _LIGHT = "1E293B", "0D9488", "F1F5F9"
_WHITE = "FFFFFF"

_TITLE_FONT = Font(bold=True, size=14, color=_WHITE)
_HEADER_FONT = Font(bold=True, size=9, color=_WHITE)
_LABEL_FONT = Font(size=9)
_BOLD_FONT = Font(bold=True, size=9)
_SUB_FONT = Font(bold=True, size=9, color=_ACCENT)
_TAG_FONT = Font(italic=True, size=8, color="475569")

_TITLE_FILL = PatternFill("solid", fgColor=_NAVY)
_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_SUB_FILL = PatternFill("solid", fgColor=_LIGHT)
_TOTAL_FILL = PatternFill("solid", fgColor=_LIGHT)

_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_RIGHT = Alignment(horizontal="right")
_LEFT = Alignment(horizontal="left")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _title(ws, text, span):
    ws.append([text])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=max(span, 1))
    c = ws.cell(ws.max_row, 1)
    c.font, c.fill, c.alignment = _TITLE_FONT, _TITLE_FILL, _LEFT
    ws.row_dimensions[ws.max_row].height = 22


def _header(ws, columns):
    ncols = len(columns)
    ws.append(["Particulars"] + [c[0] for c in columns])
    r1 = ws.max_row
    ws.append([""] + [c[1] for c in columns])
    r2 = ws.max_row
    for col in range(1, ncols + 2):
        for r, fill, font in ((r1, _HEADER_FILL, _HEADER_FONT), (r2, _SUB_FILL, _TAG_FONT)):
            cell = ws.cell(r, col)
            cell.fill, cell.font = fill, font
            cell.alignment = _CENTER if col > 1 else _LEFT
            cell.border = _BORDER


def _data_row(ws, row, ncols):
    style = row["style"]
    ws.append([row["label"]] + list(row["values"]))
    r = ws.max_row
    lab = ws.cell(r, 1)
    lab.alignment = _LEFT
    if style == "sub":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols + 1)
        lab.font, lab.fill = _SUB_FONT, _SUB_FILL
        return
    bold = style == "bold"
    lab.font = _BOLD_FONT if bold else _LABEL_FONT
    lab.border = _BORDER
    for col in range(2, ncols + 2):
        cell = ws.cell(r, col)
        cell.alignment, cell.border = _RIGHT, _BORDER
        cell.font = _BOLD_FONT if bold else _LABEL_FONT
        if row["money"] and isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0"
        if bold:
            cell.fill = _TOTAL_FILL


def _autosize(ws, ncols, first_w=42):
    ws.column_dimensions["A"].width = first_w
    for col in range(2, ncols + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _render_table(ws, sec):
    cols = sec["columns"]
    ncols = len(cols)
    _title(ws, sec["title"], ncols + 1)
    _header(ws, cols)
    for row in sec["rows"]:
        _data_row(ws, row, ncols)
    _autosize(ws, ncols)


def _render_kv(ws, sec):
    _title(ws, sec["title"], 2)
    ws.append(["Field", "Details"])
    for col in (1, 2):
        c = ws.cell(ws.max_row, col)
        c.font, c.fill = _HEADER_FONT, _HEADER_FILL
    for k, v in sec["pairs"]:
        ws.append([k, v])
        ws.cell(ws.max_row, 1).font = _BOLD_FONT
        ws.cell(ws.max_row, 2).alignment = _LEFT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42


def build_cma_workbook(results: Dict[str, Any]) -> bytes:
    """Build the CMA workbook from shared sections and return xlsx bytes."""
    wb = Workbook()
    wb.remove(wb.active)
    for sec in build_sections(results):
        ws = wb.create_sheet(sec["sheet"][:31])
        if sec["kind"] == "kv":
            _render_kv(ws, sec)
        else:
            _render_table(ws, sec)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
